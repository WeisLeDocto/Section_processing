# coding: utf-8

import numpy as np
from tools import select_folder
import tkinter as tk
from sys import exit
from openpyxl import load_workbook
from matplotlib import pyplot as plt

if __name__ == '__main__':

  # Base hidden window, necessary for using the TopLevel
  root = tk.Tk()
  root.withdraw()

  # Getting the path to working directory
  base_folder = select_folder()

  # In case the user wants to cancel
  if base_folder is None:
    root.destroy()
    exit()

  folders = [dir_ for dir_ in base_folder.iterdir()
             if list(dir_.rglob('*.xlsx'))]

  for folder in folders:

    data_files = list(folder.rglob('*.xlsx'))

    min_area = float('inf')
    max_area = 0

    for file in data_files:
      workbook = load_workbook(file)
      sheet = workbook.active
      values = np.array([*list(*sheet.iter_cols(min_col=3, max_col=3,
                                                values_only=True))][1:])
      values = values * 0.221 * 0.221
      min_area = min(min_area, np.min(values))
      max_area = max(max_area, np.max(values))

    colors = ['b', 'r', 'g']
    labels = []

    fig = plt.figure()
    ax1 = fig.add_subplot(2, 1, 1)
    ax2 = fig.add_subplot(2, 1, 2)

    for color, file in zip(colors, data_files):

      workbook = load_workbook(file)
      sheet = workbook.active
      values = np.array([*list(*sheet.iter_cols(min_col=3, max_col=3,
                                                values_only=True))][1:])
      values = values * 0.221 * 0.221
      overall_area = sheet['F2'].value * 0.221 * 0.221

      bins = np.logspace(np.log10(min_area), np.log10(max_area), 75)

      val, bins, _ = ax1.hist(values, density=False, alpha=1, linewidth=1.2,
                              edgecolor=color, bins=bins, facecolor='None',
                              histtype='step')

      val /= overall_area

      ax2.hist(bins[1:], bins[1:], weights=val, density=False, alpha=1,
               linewidth=1.2, edgecolor=color, facecolor='None',
               histtype='step')

      ax1.set_xscale('log')
      ax1.set_yscale('log')
      ax1.set_ylabel('Nr of items')

      ax2.set_xscale('log')
      ax2.set_yscale('log')
      ax2.set_ylabel('Items per square micrometer')
      labels.append(file.parent.stem)
    ax1.legend(labels)
    ax2.legend(labels)
    plt.show()
